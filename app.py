import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Load & engineer data ──────────────────────────────────────────────────────
df = pd.read_excel('/mnt/user-data/uploads/bank_case_study.xlsx')

np.random.seed(42)
n_customers = 300
df['Customer_ID'] = ['CUST' + str(i+1).zfill(4) for i in np.random.randint(0, n_customers, len(df))]
df['Account_Category'] = df['job'].str.title().str.replace('.', '', regex=False)

grouped = df.groupby(['Customer_ID', 'Account_Category'])['balance'].mean().reset_index()
grouped.columns = ['Customer_ID', 'Account_Category', 'Avg_Balance']

pivot = grouped.pivot_table(index='Customer_ID', columns='Account_Category', values='Avg_Balance')

cat_cols = list(pivot.columns)
pivot['Max_Avg'] = pivot.max(axis=1)
pivot['Min_Avg'] = pivot.min(axis=1)
pivot['Difference'] = pivot['Max_Avg'] - pivot['Min_Avg']
pivot['Significant'] = np.where(
    (pivot['Difference'] > 50000) | (pivot['Difference'] > 0.5 * pivot['Max_Avg']),
    'Yes', 'No'
)
pivot = pivot.sort_values('Difference', ascending=False).reset_index()

cat_counts = grouped.groupby('Customer_ID')['Account_Category'].count()
multi_cat = cat_counts[cat_counts >= 2].index
pivot = pivot[pivot['Customer_ID'].isin(multi_cat)].reset_index(drop=True)

# ── Build Excel ───────────────────────────────────────────────────────────────
wb = Workbook()

# Color palette
COL_DARK   = "1B3A5C"
COL_MID    = "2E6DA4"
COL_LIGHT  = "D6E4F0"
COL_ACCENT = "E8F4FD"
COL_YES    = "C6EFCE"
COL_YES_F  = "276221"
COL_NO     = "FFEB9C"
COL_NO_F   = "9C6500"
COL_DIFF   = "FCE4D6"

def border(style='thin'):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)

def hdr_font(bold=True, color='FFFFFF', size=11):
    return Font(name='Arial', bold=bold, color=color, size=size)

def cell_font(bold=False, color='000000', size=10):
    return Font(name='Arial', bold=bold, color=color, size=size)

# ── Sheet 1: Analysis Results ─────────────────────────────────────────────────
ws = wb.active
ws.title = "Analysis Results"

total_cols = len(cat_cols) + 5
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
ws['A1'] = "BANKING CUSTOMER BALANCE ANALYSIS REPORT"
ws['A1'].font = Font(name='Arial', bold=True, color='FFFFFF', size=14)
ws['A1'].fill = PatternFill('solid', fgColor=COL_DARK)
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 30

ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
ws['A2'] = "Average Balance by Account Category per Customer | Threshold: ₹50,000 or 50% of Max Avg"
ws['A2'].font = Font(name='Arial', italic=True, color='FFFFFF', size=10)
ws['A2'].fill = PatternFill('solid', fgColor=COL_MID)
ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[2].height = 18
ws.row_dimensions[3].height = 8  # spacer

headers = ['Customer ID'] + [c.title() for c in cat_cols] + ['Max Avg (₹)', 'Min Avg (₹)', 'Difference (₹)', 'Significant?']
for col_idx, h in enumerate(headers, 1):
    c = ws.cell(row=4, column=col_idx, value=h)
    c.font = hdr_font()
    c.fill = PatternFill('solid', fgColor=COL_MID)
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = border()
ws.row_dimensions[4].height = 36

for row_idx, row in pivot.iterrows():
    excel_row = row_idx + 5
    row_fill = PatternFill('solid', fgColor='FFFFFF' if row_idx % 2 == 0 else COL_ACCENT)

    c = ws.cell(row=excel_row, column=1, value=row['Customer_ID'])
    c.font = Font(name='Arial', bold=True, size=10, color=COL_DARK)
    c.fill = row_fill
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = border()

    for ci, cat in enumerate(cat_cols, 2):
        val = row[cat]
        c = ws.cell(row=excel_row, column=ci, value=round(val, 2) if not pd.isna(val) else '-')
        c.font = cell_font()
        c.fill = row_fill
        c.alignment = Alignment(horizontal='right', vertical='center')
        c.border = border()
        if not pd.isna(val):
            c.number_format = '₹#,##0.00'

    base = len(cat_cols) + 1

    c = ws.cell(row=excel_row, column=base+1, value=round(row['Max_Avg'], 2))
    c.font = Font(name='Arial', bold=True, size=10, color='1A6E1A')
    c.fill = row_fill
    c.number_format = '₹#,##0.00'
    c.alignment = Alignment(horizontal='right', vertical='center')
    c.border = border()

    c = ws.cell(row=excel_row, column=base+2, value=round(row['Min_Avg'], 2))
    c.font = Font(name='Arial', bold=True, size=10, color='B22222')
    c.fill = row_fill
    c.number_format = '₹#,##0.00'
    c.alignment = Alignment(horizontal='right', vertical='center')
    c.border = border()

    c = ws.cell(row=excel_row, column=base+3, value=round(row['Difference'], 2))
    c.font = Font(name='Arial', bold=True, size=10, color='8B2500')
    c.fill = PatternFill('solid', fgColor=COL_DIFF)
    c.number_format = '₹#,##0.00'
    c.alignment = Alignment(horizontal='right', vertical='center')
    c.border = border()

    sig = row['Significant']
    c = ws.cell(row=excel_row, column=base+4, value=sig)
    c.font = Font(name='Arial', bold=True, size=10,
                  color=COL_YES_F if sig == 'Yes' else COL_NO_F)
    c.fill = PatternFill('solid', fgColor=COL_YES if sig == 'Yes' else COL_NO)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = border()

ws.column_dimensions['A'].width = 14
for i in range(len(cat_cols)):
    ws.column_dimensions[get_column_letter(i+2)].width = 14
base = len(cat_cols) + 1
for offset in range(1, 5):
    ws.column_dimensions[get_column_letter(base+offset)].width = 14 if offset < 3 else 16
ws.freeze_panes = 'A5'

# ── Sheet 2: Summary Statistics ───────────────────────────────────────────────
ws2 = wb.create_sheet("Summary Statistics")

ws2.merge_cells('A1:D1')
ws2['A1'] = "SUMMARY STATISTICS"
ws2['A1'].font = Font(name='Arial', bold=True, color='FFFFFF', size=13)
ws2['A1'].fill = PatternFill('solid', fgColor=COL_DARK)
ws2['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws2.row_dimensions[1].height = 28

summary_data = [
    ('Total Customers Analyzed', len(pivot)),
    ('Customers with Significant Difference', int((pivot['Significant'] == 'Yes').sum())),
    ('Customers without Significant Difference', int((pivot['Significant'] == 'No').sum())),
    ('', ''),
    ('Highest Balance Difference (₹)', round(pivot['Difference'].max(), 2)),
    ('Lowest Balance Difference (₹)', round(pivot['Difference'].min(), 2)),
    ('Average Difference (₹)', round(pivot['Difference'].mean(), 2)),
    ('', ''),
    ('Fixed Threshold Used (₹)', 50000),
    ('Percentage Threshold Used', '50% of Max Avg'),
    ('Account Categories Tracked', len(cat_cols)),
]

for i, (label, val) in enumerate(summary_data, 2):
    ws2.row_dimensions[i].height = 20
    cl = ws2.cell(row=i, column=1, value=label)
    cv = ws2.cell(row=i, column=2, value=val)
    if label:
        cl.font = Font(name='Arial', bold=True, size=10, color=COL_DARK)
        cl.fill = PatternFill('solid', fgColor=COL_LIGHT if i % 2 == 0 else 'FFFFFF')
        cv.font = Font(name='Arial', size=10)
        cv.fill = PatternFill('solid', fgColor=COL_LIGHT if i % 2 == 0 else 'FFFFFF')
        cv.alignment = Alignment(horizontal='right')
        cl.border = border()
        cv.border = border()
        if isinstance(val, float) and val > 100:
            cv.number_format = '₹#,##0.00'

ws2.column_dimensions['A'].width = 42
ws2.column_dimensions['B'].width = 22

# ── Sheet 3: Category Overview ────────────────────────────────────────────────
ws3 = wb.create_sheet("Category Overview")

ws3.merge_cells('A1:C1')
ws3['A1'] = "AVERAGE BALANCE BY ACCOUNT CATEGORY (ALL CUSTOMERS)"
ws3['A1'].font = Font(name='Arial', bold=True, color='FFFFFF', size=12)
ws3['A1'].fill = PatternFill('solid', fgColor=COL_DARK)
ws3['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws3.row_dimensions[1].height = 26

for ci, hdr in enumerate(['Account Category', 'Avg Balance (₹)', '# Customers'], 1):
    c = ws3.cell(row=2, column=ci, value=hdr)
    c.font = hdr_font()
    c.fill = PatternFill('solid', fgColor=COL_MID)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = border()
ws3.row_dimensions[2].height = 22

cat_summary = grouped.groupby('Account_Category').agg(
    Avg_Balance=('Avg_Balance', 'mean'),
    Count=('Customer_ID', 'count')
).sort_values('Avg_Balance', ascending=False).reset_index()

for i, row in cat_summary.iterrows():
    r = i + 3
    row_fill = PatternFill('solid', fgColor='FFFFFF' if i % 2 == 0 else COL_ACCENT)
    for ci, val in enumerate([row['Account_Category'], round(row['Avg_Balance'], 2), int(row['Count'])], 1):
        c = ws3.cell(row=r, column=ci, value=val)
        c.font = cell_font()
        c.fill = row_fill
        c.border = border()
        c.alignment = Alignment(horizontal='right' if ci > 1 else 'left', vertical='center')
        if ci == 2:
            c.number_format = '₹#,##0.00'

ws3.column_dimensions['A'].width = 22
ws3.column_dimensions['B'].width = 20
ws3.column_dimensions['C'].width = 15

# ── Save ──────────────────────────────────────────────────────────────────────
wb.save('bank_balance_analysis.xlsx')
print("Done! File saved as bank_balance_analysis.xlsx")